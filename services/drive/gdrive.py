from __future__ import print_function

from os import path, getenv
from typing import Any, Literal, Self
from uuid import uuid4
from enum import Enum
import asyncio
from datetime import timedelta, datetime
from db import interfaces as I
from io import BytesIO
from docx import Document
from openpyxl import load_workbook

# from google.auth.transport._aiohttp_requests import Request
from google.auth.transport.requests import Request, AuthorizedSession
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build, Resource
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaIoBaseDownload
# If modifying these scopes, delete the file token.json.


class MimeTypes(Enum):

    FOLDER = "application/vnd.google-apps.folder"
    TEXT = "text/plain"
    OFFICE_DOCUMENT = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    OFFICE_SPREADSHEET = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    GOOGLE_DOCUMENT = "application/vnd.google-apps.document"
    GOOGLE_SPREADSHEET = "application/vnd.google-apps.spreadsheet"


class Extensions(Enum):

    TXT = "txt"
    DOCX = "docx"
    XLSX = "xlsx"


class Headers(Enum):
    Channel_ID = "X-Goog-Channel-ID"
    Channel_Token = "X-Goog-Channel-Token"
    Channel_Expiration = "X-Goog-Channel-Expiration"
    Resource_ID = "X-Goog-Resource-ID"
    Resource_URI = "X-Goog-Resource-URI"
    Resource_State = "X-Goog-Resource-State"
    Changed = "X-Goog-Changed"
    Message_Number = "X-Goog-Message-Number"


class ResourceStates(Enum):
    SYNC = "sync"
    ADD = "add"
    REMOVE = "remove"
    UPDATE = "update"
    TRASH = "trash"
    UNTRASH = "untrash"
    CHANGE = "change"


class GDriveAuth:

    SCOPES = [
        'https://www.googleapis.com/auth/drive',
        'https://www.googleapis.com/auth/drive.activity',
    ]

    __CREDS = path.join("services", "drive", "credentials.json")
    __TOKEN = path.join("services", "drive", "token.json")

    def __init__(self) -> None:
        self._creds = self.__init_creds()
        self.__authorize()
        self._service = self.__init_service()

    @classmethod
    def __init_creds(cls) -> Credentials:
        creds = None

        if path.exists(cls.__TOKEN):
            creds = Credentials.from_authorized_user_file(
                cls.__TOKEN, cls.SCOPES)
        return creds

    def __init_service(self) -> Resource:
        service = build('drive', 'v3', credentials=self._creds)
        return service

    def __authorize(self):
        creds = self._creds
        assert creds is self._creds
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file(
                    self.__CREDS, self.SCOPES)
                creds = flow.run_local_server(port=8080)

            with open(self.__TOKEN, 'w') as token:
                token.write(creds.to_json())


class GDrive(GDriveAuth):

    def share_dir(self, dir_id: str, permissions: Literal["owner", "organizer", "fileOrganizer", "writer", "commenter", "reader"]):
        permission = {
            'type': 'anyone',
            'role': permissions,
        }
        try:
            self._service.permissions().create(
                fileId=dir_id, body=permission, fields='id').execute()
        except HttpError as error:
            print(f'An error occurred: {error}')

    def listdir(self, dir_id: str):
        try:
            files = self._service.files().list(
                q=f"'{dir_id}' in parents").execute()

            print(files)

        except HttpError as error:
            print(f'An error occurred: {error}')

        return files

    def mkdir(self, name: str, parent_folder: str = None):
        metadata = {'name': name,
                    'mimeType': 'application/vnd.google-apps.folder',
                    'type': 'anyone',
                    'role': 'writer',
                    }
        if parent_folder is not None:
            metadata["parents"] = [parent_folder]

        try:
            file = self._service.files().create(body=metadata, fields="id",
                                                supportsAllDrives=True).execute()

        except HttpError as error:
            print(f'An error occurred: {error}')

        return file.get("id")

    def get_file_content(self, file_id: str, mime_type: MimeTypes) -> BytesIO:

        match mime_type:
            case MimeTypes.GOOGLE_DOCUMENT:
                request = self._service.files().export(
                    fileId=file_id, mimeType=MimeTypes.OFFICE_DOCUMENT.value)

            case MimeTypes.GOOGLE_SPREADSHEET:
                request = self._service.files().export(
                    fileId=file_id, mimeType=MimeTypes.OFFICE_SPREADSHEET.value)

            case _:
                request = self._service.files().get_media(fileId=file_id)

        try:
            file_content = BytesIO()
            downloader = MediaIoBaseDownload(file_content, request)

            done = False
            while not done:
                status, done = downloader.next_chunk()
                print(f"{status=}, {done=}")
        except HttpError as error:
            print(f'An error occurred: {error}')

        file_content.seek(0)
        return file_content

    def get_parent_id(self, file_id: str) -> str:
        return self._service.files().get(fileId=file_id, fields='parents').execute()["parents"][0]

    def __get_dir_id(self, dirname: str):
        try:
            folder = self._service.files().list(
                q=f"'{self.BASEDIR_ID}' in parents and name = '{dirname}'").execute()
            print("THIS IS FOLDER", folder)
        except HttpError as error:
            print(f'An error occurred: {error}')

        file = folder.get("files")[0]

        return file.get("id")


class GDriveEventsPoller:

    __SELF: "GDriveEventsPoller" = None
    __DRIVE = GDrive

    def __new__(cls) -> Self:
        if cls.__SELF is not None:
            return cls.__SELF
        return super().__new__(cls)

    def __init__(self) -> None:
        self.hanler_address = "https://babyfalcon.ru/drive/events/"
        self.start_page_token = self.__init_start_page_token()
        self.__channel = {}

    async def start_polling(self):

        # while True:
        self.register_event_handler()
#            await sleep(timedelta(weeks=1).total_seconds())

    def delete_channel(self):
        print("deleting channel", self.__channel)
        self.__DRIVE()._service.channels().stop(body=self.__channel).execute()

    def __init_start_page_token(self):
        token = self.__DRIVE()._service.changes().getStartPageToken().execute()
        return token["startPageToken"]

    def register_event_handler(self) -> None:
        drive = self.__DRIVE()
        body = {
            'id': str(uuid4()),
            'type': 'web_hook',
            'expiration': self.get_expiration_time(),
            'address': self.hanler_address,
        }

        self.__channel = drive._service.changes().watch(
            pageToken=self.start_page_token, body=body).execute()

    def get_expiration_time(self):
        datetime_expiration = datetime.now() + timedelta(weeks=1)
        return int(datetime.timestamp(datetime_expiration)*1000)


class GDriveEventProcesser:

    ROOT_DIR_ID = "0AM5SefLlkFoLUk9PVA"

    def _get_changed_files(self, changes: list[dict]) -> dict[str, I.File]:

        changes_mapping = {}
        files_mapping = {}

        for change in changes:
            file_id = change["fileId"]
            file = I.File.model_validate(change["file"])

            changes_mapping[file_id] = file
            if (not self._is_folder(file)) and (not self._is_on_root_level(file)):
                files_mapping[file_id] = file

        self._set_parents_folders(files_mapping, changes_mapping)
        return files_mapping

    def _is_folder(self, file: I.File) -> bool:
        return file.mime_type == MimeTypes.FOLDER.value

    def _is_on_root_level(self, file: I.File):
        return file.parents[0] == self.ROOT_DIR_ID

    def _set_parents_folders(self, files_mapping: dict[str, I.File], changes_mapping: dict[str, I.File]) -> None:
        drive = GDrive()
        for file in files_mapping.values():
            file_id = file.id
            parent_id = file.parents[0]
            while parent_id != self.ROOT_DIR_ID:
                file_id = parent_id
                parent_id = drive.get_parent_id(file_id)

            file.parents[0] = file_id


class GDriveContentPreprocessor:

    def get_file_content(self, drive: GDrive, file: I.File):
        mime_type = MimeTypes(file.mime_type)
        content = drive.get_file_content(file.id, mime_type)
        content_string: str

        match mime_type:
            case MimeTypes.GOOGLE_SPREADSHEET | MimeTypes.OFFICE_SPREADSHEET:
                content_string = self.process_table(content)
            case MimeTypes.GOOGLE_DOCUMENT | MimeTypes.OFFICE_DOCUMENT:
                content_string = self.process_document(content)
            case _:
                content_string = self.process_txt(content)

        file.content = content_string

    def process_table(self, content: BytesIO):
        workbook = load_workbook(content)
        sheet = workbook.active
        col_names = (cell.value for cell in sheet[1])
        col_names = tuple(filter(lambda name: name is not None, col_names))
        max_col = len(col_names)

        result = ""
        for i in range(2, sheet.max_row+1):
            for j in range(max_col):
                result += f"{col_names[j]}:\n-{sheet[i][j].value}\n"
            result += "\n"

        return result

    def process_document(self, content: BytesIO):
        document = Document(content)
        return " ".join([paragraph.text for paragraph in document.paragraphs])

    def process_txt(self, content: BytesIO):
        res = content.read().decode()
        print(f"\033[93m{type(res)}\033[0m")
        return res


class GDriveEventsHandler(GDriveContentPreprocessor, GDriveEventProcesser):

    __SELF: "GDriveEventsHandler" = None
    __DRIVE = GDrive

    def __new__(cls) -> Self:
        print("I am in new method and this is __SELF=", id(cls.__SELF))
        if cls.__SELF is None:
            self = super().__new__(cls)
            self.__task = None
            self.__resource_uris = []
            self.__events_count = 0
            cls.__SELF = self

        return cls.__SELF

    def __init__(self) -> None:
        self.__task: asyncio.Task | None
        self.__resource_uris: list
        self.__events_count: int

    def __process_event(self) -> dict[str, I.File]:
        print("\033[93mStart to proceed\033[0m")
        drive = self.__DRIVE()
        session = AuthorizedSession(drive._creds)
        changes: list[dict] = []
        

        fields = f'nextPageToken,newStartPageToken,changes(fileId,kind,removed,file(name,mimeType,parents,id,description,trashed,webContentLink,fileExtension))'
        for uri in self.__resource_uris:
            resp = session.get(f"{uri}&fields={fields}")
            print(resp)
            changes += resp.json()["changes"]

        changes = changes[-self.__events_count:]

        # Resetting for next events    
        self.__events_count = 0
        self.__resource_uris.clear()


        print(f"2{changes=}")
        files_mapping = self._get_changed_files(changes)
        for file in files_mapping.values():
            self.get_file_content(drive, file)
        
        return files_mapping

    async def __handle_event(self):
        await asyncio.sleep(5)
        return self.__process_event()

    async def handle_event(self, headers: dict) -> dict[str, I.File] | None:

        if headers[Headers.Resource_State.value] == ResourceStates.SYNC.value:
            return
        print("Resource uris", self.__resource_uris)
        if self.__task is not None:
            print("\033[93mCanceling task\033[0m")
            self.__task.cancel()

        self.__events_count += 1
        self.__task = asyncio.create_task(self.__handle_event())

        self._add_resource_uri(headers[Headers.Resource_URI.value])
        task_res = await self.__task
        return task_res

    def _add_resource_uri(self, uri: str):
        if uri not in self.__resource_uris:
            self.__resource_uris.append(uri)


class GDriveEventsManager:

    def __init__(self) -> None:
        self.POLLER = GDriveEventsPoller()
        self.HANDLER = GDriveEventsHandler()
